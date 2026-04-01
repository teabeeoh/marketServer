#!/usr/bin/env python3
"""
Excel Template Filler Script

This script retrieves financial data for a given Yahoo Finance ticker symbol
and fills an Excel template with the data.

Usage:
    python fill_excel_template.py <TICKER_SYMBOL> [output_file.xlsx]

Example:
    python fill_excel_template.py AAPL
    python fill_excel_template.py SAP.DE my_output.xlsx
"""

import sys
import argparse
from datetime import date
import pandas as pd
from pathlib import Path
from copy import copy
from openpyxl import load_workbook
from financial_data_service import fetch_financial_data, fetch_company_info


def fill_excel_template(ticker_symbol: str, template_path: Path, output_path: Path):
    """
    Fill Excel template with financial data for the given ticker symbol.

    The template file is NEVER modified - it is only read and copied.

    Args:
        ticker_symbol: Yahoo Finance ticker symbol (e.g., 'AAPL', 'SAP.DE')
        template_path: Path to the Excel template file (read-only)
        output_path: Path where the filled Excel file will be saved

    Raises:
        ValueError: If ticker is invalid or data cannot be fetched
        FileNotFoundError: If the template file does not exist
    """
    print(f"Fetching financial data for {ticker_symbol}...")

    df = fetch_financial_data(ticker_symbol)
    company_info = fetch_company_info(ticker_symbol)

    # TODO: Check if removing this results in more data or has any effect?
    # Limit to last 5 years
    df = df.head(5)

    print(f"Retrieved data for {len(df)} years")

    # Ensure we never overwrite the template
    if output_path.resolve() == template_path.resolve():
        raise ValueError(f"Output path cannot be the same as template path: {template_path}")

    # Load the Excel template (READ-ONLY - will never be modified)
    print(f"Loading template from {template_path}...")
    template_workbook = load_workbook(template_path)

    # Get the first (template) sheet
    template_sheet = template_workbook.active

    # Create a NEW workbook for output
    from openpyxl import Workbook
    output_workbook = Workbook()
    output_workbook.remove(output_workbook.active)  # Remove default sheet

    # Copy the template sheet to the new workbook
    new_sheet_name = ticker_symbol.replace('.', '_')  # Excel doesn't like dots in sheet names
    print(f"Creating new workbook with sheet '{new_sheet_name}' based on template...")

    # Copy worksheet from template to output workbook
    output_sheet = output_workbook.create_sheet(title=new_sheet_name)

    # Copy all cells from template sheet to new sheet
    for row in template_sheet.iter_rows():
        for cell in row:
            new_cell = output_sheet[cell.coordinate]
            new_cell.value = cell.value
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = cell.number_format
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

    # Copy column dimensions
    for col_letter, dimension in template_sheet.column_dimensions.items():
        output_sheet.column_dimensions[col_letter].width = dimension.width

    # Copy row dimensions
    for row_num, dimension in template_sheet.row_dimensions.items():
        output_sheet.row_dimensions[row_num].height = dimension.height

    # Copy merged cells
    for merged_cell_range in template_sheet.merged_cells.ranges:
        output_sheet.merge_cells(str(merged_cell_range))

    # Close template workbook (we're done reading it)
    template_workbook.close()

    # Fill the new sheet with data
    print("Filling sheet with financial data...")

    # The template has:
    # - Row 1-4: Metadata (Name, ISIN, Sektor, Währung, Dashboard)
    # - Row 5: Headers (Jahr, Umsatz, J, FCF, Dividende, etc.)
    # - Row 6+: Data rows
    # - Row 22: DCF assumptions (CAGR, avg margin, current price)

    # --- Header section (row 2) ---
    output_sheet['A2'] = company_info['name']
    output_sheet['C2'] = company_info['sector']
    output_sheet['D2'] = company_info['currency']
    output_sheet['L2'] = date.today().strftime('%d.%m.%Y')

    # --- Data rows (row 6+) ---
    # Column mapping based on template structure (Row 5 headers)
    column_mapping = {
        'Year': 'A',
        'Total Revenue (mn)': 'B',
        'Net Income Common Stockholders (mn)': 'C',
        'Free Cash Flow (mn)': 'D',
        'Dividend per Share': 'E',
        'Ordinary Shares Number (mn)': 'F',
        'Stockholders Equity (mn)': 'G',
        'Total Assets (mn)': 'H',
        'Goodwill (mn)': 'I',
        'Other Intangible Assets (mn)': 'J'
    }

    start_row = 6
    for row_idx, (_, row_data) in enumerate(df.iterrows(), start=start_row):
        for col_name, excel_col in column_mapping.items():
            cell = f"{excel_col}{row_idx}"
            value = row_data[col_name]

            if col_name == 'Year' and isinstance(value, str) and len(value) == 10:
                # Reformat YYYY-MM-DD → DD.MM.YYYY
                try:
                    value = f"{value[8:10]}.{value[5:7]}.{value[0:4]}"
                except Exception:
                    pass

            # Handle None/NaN values
            if value is None or (isinstance(value, float) and pd.isna(value)):
                output_sheet[cell] = ""
            else:
                output_sheet[cell] = value

    # --- DCF assumptions (row 22) ---
    revenues = [r for r in df['Total Revenue (mn)'] if r is not None and not (isinstance(r, float) and pd.isna(r))]
    if len(revenues) >= 2:
        n = len(revenues)
        cagr = (revenues[0] / revenues[n - 1]) ** (1 / (n - 1)) - 1
        output_sheet['C22'] = cagr

    margins = []
    for _, row in df.iterrows():
        rev = row['Total Revenue (mn)']
        ni = row['Net Income Common Stockholders (mn)']
        if (rev is not None and ni is not None and
                not (isinstance(rev, float) and pd.isna(rev)) and
                not (isinstance(ni, float) and pd.isna(ni)) and
                rev != 0):
            margins.append(ni / rev)
    if margins:
        output_sheet['E22'] = sum(margins) / len(margins)

    if company_info.get('current_price') is not None:
        output_sheet['F22'] = company_info['current_price']

    # Save the NEW workbook (template remains untouched)
    print(f"Saving workbook to {output_path}...")
    output_workbook.save(output_path)
    output_workbook.close()
    print(f"✓ Successfully created {output_path} with financial data for {ticker_symbol}")
    print(f"✓ Template file {template_path} was NOT modified")


def main():
    """Main entry point for the script."""
    parser = argparse.ArgumentParser(
        description='Fill Excel template with Yahoo Finance financial data',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s AAPL
  %(prog)s SAP.DE
  %(prog)s MSFT output.xlsx
        """
    )

    parser.add_argument(
        'ticker',
        help='Yahoo Finance ticker symbol (e.g., AAPL, MSFT, SAP.DE)'
    )

    parser.add_argument(
        'output',
        nargs='?',
        help='Output Excel file path (default: <ticker>.xlsx)'
    )

    parser.add_argument(
        '--template',
        default='resources/excel_template.xlsx',
        help='Path to Excel template (default: resources/excel_template.xlsx)'
    )

    args = parser.parse_args()

    # Determine paths
    # Load template relative to current working directory (execution directory)
    cwd = Path.cwd()
    template_path = (cwd / args.template).resolve()

    # Determine output file name
    if args.output:
        output_path = Path(args.output).resolve()
    else:
        safe_ticker = args.ticker.replace('.', '_')
        output_path = (cwd / f"out/{safe_ticker}.xlsx").resolve()

    # Fill the template
    try:
        fill_excel_template(args.ticker, template_path, output_path)
    except (ValueError, FileNotFoundError) as e:
        print(f"Error: {e}")
        sys.exit(1)


if __name__ == '__main__':
    main()
